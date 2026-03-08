from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
import os
import json
import tempfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "cambia_esta_clave_en_produccion")
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get("DATABASE_URL", "sqlite:///datos.db")
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ── Credenciales admin ────────────────────────────────────────────────────────
ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "seguridad2024")

# ── Config notificaciones Gmail ───────────────────────────────────────────────
GMAIL_USER   = os.environ.get("GMAIL_USER", "")
GMAIL_PASS   = os.environ.get("GMAIL_PASS", "")
NOTIFY_EMAIL = os.environ.get("NOTIFY_EMAIL", "fgarollaaero@gmail.com")


def enviar_notificacion(sector, tipo, area):
    """Envía mail via SMTP con TLS (puerto 587) — compatible con Railway."""
    if not GMAIL_USER or not GMAIL_PASS:
        return
    try:
        msg = MIMEMultipart()
        msg["From"]    = GMAIL_USER
        msg["To"]      = NOTIFY_EMAIL
        msg["Subject"] = f"[SMS Aeronautico] Nuevo reporte - {sector.capitalize()}"

        cuerpo = (
            f"Se recibio un nuevo reporte en el Sistema de Gestion de Seguridad.\n\n"
            f"  Sector: {sector.capitalize()}\n"
            f"  Tipo:   {tipo.capitalize()}\n"
            f"  Area:   {area}\n\n"
            f"Ingresa al panel administrativo para ver el detalle completo."
        )
        msg.attach(MIMEText(cuerpo, "plain", "utf-8"))

        # Puerto 587 con STARTTLS — Railway permite este puerto
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=10) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()
            smtp.login(GMAIL_USER, GMAIL_PASS)
            smtp.sendmail(GMAIL_USER, NOTIFY_EMAIL, msg.as_string())
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")


# ── Modelos ───────────────────────────────────────────────────────────────────
class Reporte(db.Model):
    id            = db.Column(db.Integer, primary_key=True)
    sector        = db.Column(db.String(20))          # 'mantenimiento' | 'operaciones'
    tipo_reporte  = db.Column(db.String(20))          # 'obligatorio' | 'voluntario'
    nombre        = db.Column(db.String(100))
    area          = db.Column(db.String(100))
    riesgo        = db.Column(db.String(200))
    fecha         = db.Column(db.String(50))
    descripcion   = db.Column(db.String(500))
    acciones      = db.Column(db.String(500))
    aeronave      = db.Column(db.String(100))
    componente    = db.Column(db.String(100))
    # Matriz OACI
    probabilidad  = db.Column(db.Integer)             # 1-5
    severidad     = db.Column(db.Integer)             # 1-5
    indice_riesgo = db.Column(db.Integer)             # prob * sev
    nivel_riesgo  = db.Column(db.String(20))          # Aceptable/Tolerable/Inaceptable
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)


# ── Helpers ───────────────────────────────────────────────────────────────────
def calcular_nivel_riesgo(prob, sev):
    indice = prob * sev
    if indice <= 4:
        return indice, "Aceptable"
    elif indice <= 9:
        return indice, "Tolerable"
    else:
        return indice, "Inaceptable"


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


# ── Rutas públicas ────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/reporte/<sector>", methods=["GET", "POST"])
def reporte(sector):
    if sector not in ("mantenimiento", "operaciones"):
        return redirect(url_for("index"))

    if request.method == "POST":
        tipo       = request.form.get("tipo_reporte")
        nombre     = request.form.get("nombre", "").strip()
        area       = request.form.get("area", "")
        riesgo     = request.form.get("riesgo", "")
        fecha      = request.form.get("fecha", "")
        desc       = request.form.get("descripcion", "")
        acciones   = request.form.get("acciones", "")
        aeronave   = request.form.get("aeronave", "")
        componente = request.form.get("componente", "")

        # Validaciones
        if tipo == "obligatorio" and not nombre:
            flash("El nombre es obligatorio para reportes obligatorios.", "error")
            return redirect(url_for("reporte", sector=sector))
        if not area or not desc:
            flash("El área y la descripción son campos requeridos.", "error")
            return redirect(url_for("reporte", sector=sector))

        # OACI se evalúa en el panel admin; se guarda pendiente por defecto
        nuevo = Reporte(
            sector=sector, tipo_reporte=tipo, nombre=nombre,
            area=area, riesgo=riesgo, fecha=fecha,
            descripcion=desc, acciones=acciones,
            aeronave=aeronave, componente=componente,
            probabilidad=None, severidad=None,
            indice_riesgo=None, nivel_riesgo="Pendiente"
        )
        db.session.add(nuevo)
        db.session.commit()

        enviar_notificacion(sector, tipo, area)

        flash("✅ Reporte enviado exitosamente.", "success")
        return redirect(url_for("reporte", sector=sector))

    return render_template("reporte.html", sector=sector)


# ── Rutas admin ───────────────────────────────────────────────────────────────
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        if (request.form.get("usuario") == ADMIN_USER and
                request.form.get("password") == ADMIN_PASS):
            session["admin_logged_in"] = True
            return redirect(url_for("dashboard"))
        flash("Credenciales incorrectas.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    sector  = request.args.get("sector", "todos")
    nivel   = request.args.get("nivel", "todos")
    tipo    = request.args.get("tipo", "todos")

    query = Reporte.query
    if sector != "todos":
        query = query.filter_by(sector=sector)
    if nivel != "todos":
        query = query.filter_by(nivel_riesgo=nivel)
    if tipo != "todos":
        query = query.filter_by(tipo_reporte=tipo)

    reportes = query.order_by(Reporte.created_at.desc()).all()

    # ── Estadísticas para el panel ─────────────────────────────────────────
    total        = Reporte.query.count()
    inaceptables = Reporte.query.filter_by(nivel_riesgo="Inaceptable").count()
    tolerables   = Reporte.query.filter_by(nivel_riesgo="Tolerable").count()
    aceptables   = Reporte.query.filter_by(nivel_riesgo="Aceptable").count()
    pendientes   = Reporte.query.filter_by(nivel_riesgo="Pendiente").count()
    mant_count   = Reporte.query.filter_by(sector="mantenimiento").count()
    ops_count    = Reporte.query.filter_by(sector="operaciones").count()

    # Frecuencia por área (top 5)
    from sqlalchemy import func
    areas_freq = (db.session.query(Reporte.area, func.count(Reporte.id).label("cnt"))
                  .group_by(Reporte.area).order_by(func.count(Reporte.id).desc()).limit(5).all())

    # Riesgos más reportados (top 5)
    riesgos_freq = (db.session.query(Reporte.riesgo, func.count(Reporte.id).label("cnt"))
                    .filter(Reporte.riesgo != "")
                    .group_by(Reporte.riesgo).order_by(func.count(Reporte.id).desc()).limit(5).all())

    # Tendencia últimos 6 meses
    tendencia = (db.session.query(
        func.strftime('%Y-%m', Reporte.created_at).label("mes"),
        func.count(Reporte.id).label("cnt"))
        .group_by("mes").order_by("mes").limit(6).all())

    stats = {
        "total": total,
        "inaceptables": inaceptables,
        "tolerables": tolerables,
        "aceptables": aceptables,
        "pendientes": pendientes,
        "mant": mant_count,
        "ops": ops_count,
        "areas_freq": areas_freq,
        "riesgos_freq": riesgos_freq,
        "tendencia": tendencia,
    }

    return render_template("dashboard.html", reportes=reportes, stats=stats,
                           filtro_sector=sector, filtro_nivel=nivel, filtro_tipo=tipo)


@app.route("/reporte/<int:reporte_id>/oaci", methods=["POST"])
@login_required
def evaluar_oaci(reporte_id):
    r = Reporte.query.get_or_404(reporte_id)
    try:
        prob = int(request.form.get("probabilidad"))
        sev  = int(request.form.get("severidad_oaci"))
        indice, nivel = calcular_nivel_riesgo(prob, sev)
        r.probabilidad  = prob
        r.severidad     = sev
        r.indice_riesgo = indice
        r.nivel_riesgo  = nivel
        db.session.commit()
    except Exception as e:
        flash(f"Error al guardar evaluación: {e}", "error")
    return redirect(url_for("dashboard") + "#reporte-" + str(reporte_id))


@app.route("/export")
@login_required
def export():
    reportes = Reporte.query.order_by(Reporte.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reportes SMS"

    headers = ["ID","Sector","Tipo","Nombre","Área","Riesgo","Fecha Evento",
               "Descripción","Acciones Inmediatas","Aeronave","Componente",
               "Probabilidad (OACI)","Severidad (OACI)","Índice de Riesgo",
               "Nivel de Riesgo","Fecha Registro"]

    header_fill = PatternFill("solid", fgColor="1560BD")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(len(h) + 4, 14)

    for r in reportes:
        ws.append([
            r.id, r.sector, r.tipo_reporte, r.nombre or "Anónimo",
            r.area, r.riesgo, r.fecha, r.descripcion, r.acciones,
            r.aeronave, r.componente, r.probabilidad, r.severidad,
            r.indice_riesgo, r.nivel_riesgo,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""
        ])

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        return send_file(tmp.name,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name="reportes_sms.xlsx")


# Crear tablas siempre al iniciar (funciona con gunicorn y python app.py)
with app.app_context():
    db.create_all()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
